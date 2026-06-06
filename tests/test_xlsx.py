import openpyxl
import pytest

from src.analizer import _gerar_xlsx


ASTEROIDES_MOCK = [
    {
        "id": "1",
        "nome": "Asteroide-Teste",
        "data_aprox": "2026-01-01",
        "diametro_min_m": 100.0,
        "diametro_max_m": 200.0,
        "velocidade_kmh": 50_000.0,
        "distancia_km": 1_000_000.0,
        "perigoso": 0,
        "score_risco": 42.0,
        "anomaly_score": None,
    }
]

RELATORIO_MOCK = {
    "gerado_em": "2026-01-01",
    "total_asteroides": 1,
    "potencialmente_perigosos": 0,
    "anomalias_detectadas": 0,
    "top_5_risco": ASTEROIDES_MOCK,
}


def test_xlsx_gerado_com_3_abas(tmp_path, monkeypatch):
    """_gerar_xlsx cria arquivo com exatamente 3 abas: Asteroides, Top 5 Risco, Métricas."""
    from src import config as cfg
    monkeypatch.setattr(cfg, "REPORTS_DIR", tmp_path)

    _gerar_xlsx(ASTEROIDES_MOCK, RELATORIO_MOCK)

    xlsx_files = list(tmp_path.glob("*.xlsx"))
    assert len(xlsx_files) == 1, "Deve gerar exatamente 1 arquivo XLSX"

    wb = openpyxl.load_workbook(xlsx_files[0])
    assert set(wb.sheetnames) == {"Asteroides", "Top 5 Risco", "Métricas"}


def test_xlsx_cabecalho_em_negrito(tmp_path, monkeypatch):
    """Primeira linha de cada aba deve estar em negrito."""
    from src import config as cfg
    monkeypatch.setattr(cfg, "REPORTS_DIR", tmp_path)

    _gerar_xlsx(ASTEROIDES_MOCK, RELATORIO_MOCK)

    wb = openpyxl.load_workbook(list(tmp_path.glob("*.xlsx"))[0])
    for aba in wb.sheetnames:
        ws = wb[aba]
        for cell in ws[1]:
            if cell.value is not None:
                assert cell.font.bold, f"Cabeçalho '{cell.value}' não está em negrito na aba '{aba}'"
